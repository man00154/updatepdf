import os
import re
import warnings
from pathlib import Path

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
import openpyxl
from openai import OpenAI as _OpenAI

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Sify DC – Capacity Excel Query",
    page_icon="🏢",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# THEME CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
NAVY  = "#0a0e1a"
DARK1 = "#0f1628"
DARK2 = "#141c35"
CARD  = "#1a2340"
BORD  = "#2a3a6a"
BLUE  = "#2a5298"
LBLUE = "#3d72d9"
CYAN  = "#00d4ff"
TEXT  = "#c9d8f0"
MUTED = "#7a92c0"
WHITE = "#ffffff"
GREEN = "#00c986"
AMBER = "#f5a623"
RED   = "#ff4d6d"

st.markdown(f"""
<style>
html,body,[class*="css"]{{background:{NAVY};color:{TEXT};font-family:'Segoe UI',sans-serif}}
.stApp{{background:{NAVY}}}
section[data-testid="stSidebar"]{{background:{DARK1};border-right:1px solid {BORD}}}
section[data-testid="stSidebar"] *{{color:{TEXT}!important}}
.stTabs [data-baseweb="tab-list"]{{background:{DARK1};border-radius:12px;padding:6px;gap:4px;border:1px solid {BORD}}}
.stTabs [data-baseweb="tab"]{{background:transparent;color:{MUTED};border-radius:8px;padding:8px 18px;font-weight:600;font-size:.88rem;border:none}}
.stTabs [aria-selected="true"]{{background:{BLUE};color:{WHITE}!important;box-shadow:0 2px 12px rgba(42,82,152,.5)}}
.stTabs [data-baseweb="tab-panel"]{{background:transparent;padding-top:1.2rem}}
.stSelectbox>div>div,.stMultiSelect>div>div{{background:{DARK2};border:1px solid {BORD};border-radius:8px;color:{TEXT}}}
.stTextInput>div>div{{background:{DARK2};border:1px solid {BORD};border-radius:8px}}
.stTextInput input{{color:{TEXT}}}
.stButton>button{{background:linear-gradient(135deg,{BLUE},{LBLUE});color:{WHITE};border:none;border-radius:8px;padding:8px 22px;font-weight:700;font-size:.9rem;box-shadow:0 4px 14px rgba(42,82,152,.4);transition:all .2s}}
.stButton>button:hover{{transform:translateY(-2px);box-shadow:0 6px 20px rgba(42,82,152,.6)}}
.stDataFrame{{border:1px solid {BORD};border-radius:10px;overflow:hidden}}
.stDataFrame table{{background:{DARK2};color:{TEXT}}}
.stDataFrame th{{background:{BLUE};color:{WHITE};padding:10px}}
.stDataFrame td{{border-bottom:1px solid {BORD};padding:8px 12px}}
[data-testid="metric-container"]{{background:{CARD};border:1px solid {BORD};border-radius:12px;padding:16px 20px;box-shadow:0 4px 16px rgba(0,0,0,.3)}}
[data-testid="metric-container"] label{{color:{MUTED};font-size:.82rem;font-weight:600;letter-spacing:.04em;text-transform:uppercase}}
[data-testid="metric-container"] [data-testid="stMetricValue"]{{color:{CYAN};font-size:1.7rem;font-weight:800}}
.kpi-card{{background:{CARD};border:1px solid {BORD};border-radius:14px;padding:22px 24px;box-shadow:0 4px 20px rgba(0,0,0,.35);text-align:center;height:100%}}
.kpi-val{{font-size:2rem;font-weight:900;line-height:1.1}}
.kpi-label{{font-size:.78rem;color:{MUTED};font-weight:700;text-transform:uppercase;letter-spacing:.06em;margin-top:6px}}
.kpi-sub{{font-size:.76rem;color:{GREEN};margin-top:4px}}
.section-title{{font-size:1.2rem;font-weight:800;color:{WHITE};border-left:4px solid {CYAN};padding-left:14px;margin:24px 0 16px;letter-spacing:.02em}}
.badge{{display:inline-block;background:{BLUE};color:{WHITE};border-radius:20px;padding:2px 12px;font-size:.76rem;font-weight:700;margin:2px}}
.result-box{{background:{DARK2};border:1px solid {BORD};border-radius:10px;padding:18px 22px;margin:10px 0;font-size:1rem;color:{TEXT}}}
.result-big{{font-size:2.6rem;font-weight:900;background:linear-gradient(90deg,{CYAN},{LBLUE});-webkit-background-clip:text;-webkit-text-fill-color:transparent}}
.hero{{background:linear-gradient(135deg,{DARK2} 0%,{DARK1} 50%,{NAVY} 100%);border:1px solid {BORD};border-radius:16px;padding:28px 36px;margin-bottom:24px;box-shadow:0 8px 32px rgba(0,0,0,.4)}}
.hero h1{{font-size:2rem;font-weight:900;color:{WHITE};margin:0;letter-spacing:.02em}}
.hero p{{color:{WHITE};margin:6px 0 0;font-size:.95rem}}
::-webkit-scrollbar{{width:6px;height:6px}}
::-webkit-scrollbar-track{{background:{DARK1}}}
::-webkit-scrollbar-thumb{{background:{BORD};border-radius:3px}}
::-webkit-scrollbar-thumb:hover{{background:{BLUE}}}
header[data-testid="stHeader"]{{display:none!important;visibility:hidden!important;height:0!important}}
#MainMenu{{display:none!important}}
footer{{display:none!important}}
.stDeployButton{{display:none!important}}
[data-testid="stToolbar"]{{display:none!important}}
.viewerBadge_container__1QSob{{display:none!important}}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL LOADING ENGINE
# ─────────────────────────────────────────────────────────────────────────────
_BASE = Path(__file__).parent


def _excel_dirs():
    candidates = [
        _BASE / "excel_files",
        _BASE / "attached_assets",
        _BASE,
    ]
    return [d for d in candidates if d.is_dir()]


SECTION_MARKERS = {
    "Billing Model", "Space", "Power Capacity", "Power Usage",
    "Seating Space", "Revenue", "DEMARC", "RHS", "SHS",
    "ONSITE TAPE ROTATION", "OFFSITE TAPE ROTATION",
    "SAFE VAULT", "STORE SPACE", "Contract Information",
    "Floor / Module", "Customer Name",
}

HEADER_INDICATORS = {
    "customer name", "floor", "sr. no", "sno", "s.no",
    "customer", "subscription", "caged", "uncaged", "uom", "in use",
    "power subscription", "billing model", "subscription mode",
    "ownership", "per unit rate", "mrc", "description",
}


def _is_section_row(vals) -> bool:
    non = [str(v).strip() for v in vals if v and str(v).strip() not in ("", "None")]
    if not non:
        return False
    hits = sum(1 for v in non if v in SECTION_MARKERS)
    return hits / len(non) > 0.30


def _is_header_row(vals) -> bool:
    non = [str(v).strip().lower() for v in vals if v and str(v).strip() not in ("", "None")]
    if not non:
        return False
    hits = sum(1 for v in non if any(ind in v for ind in HEADER_INDICATORS))
    return hits / len(non) >= 0.20


def _actual_col_count(rows: list) -> int:
    last = 0
    for row in rows:
        for i in range(len(row) - 1, -1, -1):
            if row[i] is not None and str(row[i]).strip() not in ("", "None"):
                last = max(last, i + 1)
                break
    return last


def _detect_header(raw_rows):
    def rs(r):
        return [str(v).strip() if v is not None else "" for v in r]

    rows = [rs(r) for r in raw_rows[:10]]

    for i in range(min(5, len(rows) - 1)):
        r1 = rows[i]
        r2 = rows[i + 1] if i + 1 < len(rows) else []
        if r2 and (_is_section_row(r1) or _is_header_row(r1)) and _is_header_row(r2):
            data_start = i + 3
            return data_start, r1, r2

    for i in range(min(5, len(rows))):
        r = rows[i]
        if _is_header_row(r):
            data_start = i + 2
            return data_start, [""] * len(r), r

    if rows:
        return 2, [""] * len(rows[0]), rows[0]
    return None, None, None


def _build_cols(g_row: list, c_row: list) -> list:
    cur_g = ""
    cols = []
    seen: dict = {}
    for g, c in zip(g_row, c_row):
        g_s = str(g).strip() if g else ""
        c_s = str(c).strip() if c else ""
        if g_s and g_s not in ("None", ""):
            cur_g = g_s
        label = c_s if c_s and c_s not in ("None", "") else ""
        raw = f"{cur_g} | {label}" if (cur_g and label) else (label or cur_g or "_col")
        cnt = seen.get(raw, 0)
        seen[raw] = cnt + 1
        cols.append(raw if cnt == 0 else f"{raw}.{cnt}")
    return cols


def _clean_df(df: pd.DataFrame) -> "pd.DataFrame | None":
    df = df.dropna(axis=1, how="all")
    df = df[df.apply(
        lambda r: any(str(v).strip() not in ("", "None", "nan") for v in r), axis=1
    )]
    try:
        err_mask = df.apply(
            lambda r: r.astype(str).str.contains(
                r"#DIV|#REF|#N/A|#VALUE", regex=True, na=False
            ).all(), axis=1
        )
        df = df[~err_mask]
    except Exception:
        pass
    for col in df.columns:
        try:
            conv = pd.to_numeric(df[col], errors="coerce")
            col_lower = col.lower()
            is_name_col = any(kw in col_lower for kw in (
                "name", "customer", "floor", "module", "model", "mode",
                "ownership", "caged", "uom", "remarks", "description",
            ))
            if not is_name_col and conv.notna().sum() / max(len(df), 1) > 0.50:
                df[col] = conv
        except Exception:
            pass
    return df.reset_index(drop=True) if len(df) >= 1 else None


def _load_ws(ws) -> "pd.DataFrame | None":
    if ws.max_row < 2:
        return None
    sample = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10), values_only=True))
    actual_ncols = _actual_col_count(sample)
    if actual_ncols < 2:
        return None
    sample = [row[:actual_ncols] for row in sample]
    hdr_start, g_row, c_row = _detect_header(sample)
    if hdr_start is None:
        return None
    g_row = list(g_row)[:actual_ncols]
    c_row = list(c_row)[:actual_ncols]
    while len(g_row) < actual_ncols:
        g_row.append("")
    while len(c_row) < actual_ncols:
        c_row.append("")
    cols = _build_cols(g_row, c_row)
    data = []
    for row in ws.iter_rows(min_row=hdr_start, max_col=actual_ncols, values_only=True):
        vals = list(row)[:len(cols)]
        while len(vals) < len(cols):
            vals.append(None)
        if any(v is not None and str(v).strip() not in ("", "None") for v in vals):
            data.append(vals)
    if not data:
        return None
    return _clean_df(pd.DataFrame(data, columns=cols))


def _load_xls_ws(sheet) -> "pd.DataFrame | None":
    try:
        import xlrd
        nrows, ncols = sheet.nrows, sheet.ncols
        if nrows < 2 or ncols < 2:
            return None

        def cv(r, c):
            try:
                cell = sheet.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    return None
                v = cell.value
                if cell.ctype == xlrd.XL_CELL_NUMBER:
                    iv = int(v)
                    return iv if iv == v else v
                return str(v).strip() if v is not None else None
            except Exception:
                return None

        sample = [[cv(ri, ci) for ci in range(ncols)] for ri in range(min(nrows, 10))]
        actual_ncols = _actual_col_count(sample)
        if actual_ncols < 2:
            return None
        sample = [row[:actual_ncols] for row in sample]
        hdr_start, g_row, c_row = _detect_header(sample)
        if hdr_start is None:
            return None
        g_row = list(g_row)[:actual_ncols]
        c_row = list(c_row)[:actual_ncols]
        while len(g_row) < actual_ncols:
            g_row.append("")
        while len(c_row) < actual_ncols:
            c_row.append("")
        cols = _build_cols(g_row, c_row)
        data = []
        for ri in range(hdr_start - 1, nrows):
            vals = [cv(ri, ci) for ci in range(actual_ncols)][:len(cols)]
            while len(vals) < len(cols):
                vals.append(None)
            if any(v is not None and str(v).strip() not in ("", "None") for v in vals):
                data.append(vals)
        if not data:
            return None
        return _clean_df(pd.DataFrame(data, columns=cols))
    except Exception:
        return None


def _label(fpath: Path) -> str:
    s = re.sub(r"Customer_and_Capacity_Tracker_", "", fpath.stem, flags=re.I)
    s = re.sub(r"_\d{10,}$", "", s)
    s = re.sub(r"_\d{2}[A-Za-z]{3}\d{2,4}", "", s)
    s = re.sub(r"[_]+", " ", s).strip()
    s = re.sub(r"\(\s*(\d+)\s*\)", r"(\1)", s).strip()
    return s if s else fpath.stem


def _file_timestamp(fpath: Path) -> int:
    m = re.search(r"_(\d{10,})(?:\.[^.]+)?$", fpath.stem)
    return int(m.group(1)) if m else 0


@st.cache_data(show_spinner=False)
def load_all() -> dict:
    """Load all 10 Excel files with all their sheets into a nested dict."""
    dirs = _excel_dirs()

    label_map: dict = {}

    def _register(fpath: Path, kind: str) -> None:
        base = _label(fpath)
        ts = _file_timestamp(fpath)
        existing = label_map.get(base)
        if existing is None or ts > existing[0]:
            label_map[base] = (ts, fpath, kind)

    seen_names: set = set()
    for d in dirs:
        for fpath in sorted(d.glob("*.xlsx")):
            if fpath.name not in seen_names:
                seen_names.add(fpath.name)
                _register(fpath, "xlsx")
        for fpath in sorted(d.glob("*.xls")):
            if fpath.suffix.lower() == ".xls" and fpath.name not in seen_names:
                seen_names.add(fpath.name)
                _register(fpath, "xls")

    result: dict = {}
    for label, (_, fpath, kind) in sorted(label_map.items()):
        if kind == "xlsx":
            try:
                wb = openpyxl.load_workbook(str(fpath), data_only=True, read_only=False, keep_links=False)
            except Exception:
                try:
                    wb = openpyxl.load_workbook(str(fpath), data_only=False, keep_links=False)
                except Exception:
                    continue
            sheets: dict = {}
            for sn in wb.sheetnames:
                try:
                    df = _load_ws(wb[sn])
                    if df is not None and len(df) >= 1:
                        sheets[sn] = df
                except Exception:
                    pass
            wb.close()
            if sheets:
                result[label] = sheets
        else:
            try:
                import xlrd
                wb = xlrd.open_workbook(str(fpath))
                sheets: dict = {}
                for sn in wb.sheet_names():
                    try:
                        df = _load_xls_ws(wb.sheet_by_name(sn))
                        if df is not None and len(df) >= 1:
                            sheets[sn] = df
                    except Exception:
                        pass
                if sheets:
                    result[label] = sheets
            except Exception:
                continue

    return result


# ─────────────────────────────────────────────────────────────────────────────
# DATA HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def combined_df(data: dict, loc_filter=None, sheet_filter=None) -> pd.DataFrame:
    frames = []
    for loc, sheets in data.items():
        if loc_filter and loc not in loc_filter:
            continue
        for sn, df in sheets.items():
            if sheet_filter and sn not in sheet_filter:
                continue
            tmp = df.copy()
            tmp.insert(0, "_Sheet", sn)
            tmp.insert(0, "_Location", loc)
            frames.append(tmp)
    if not frames:
        return pd.DataFrame()
    combined = pd.concat(frames, ignore_index=True, sort=False)
    return combined.reset_index(drop=True)


def num_cols(df: pd.DataFrame) -> list:
    return [c for c in df.columns
            if pd.api.types.is_numeric_dtype(df[c]) and not c.startswith("_")]


def txt_cols(df: pd.DataFrame) -> list:
    return [c for c in df.columns
            if not pd.api.types.is_numeric_dtype(df[c]) and not c.startswith("_")]


def find_col(df: pd.DataFrame, *pats) -> "str | None":
    for p in pats:
        for c in df.columns:
            if re.search(p, c, re.I):
                return c
    return None


def fmt(n):
    if n is None:
        return "–"
    try:
        n = float(n)
        if abs(n) >= 1_000_000:
            return f"{n / 1_000_000:.2f} M"
        if abs(n) >= 1_000:
            return f"{n / 1_000:.1f} K"
        return f"{n:,.2f}"
    except Exception:
        return str(n)


def kpi_html(value, label, sub="", color=CYAN):
    sub_html = f'<div class="kpi-sub">{sub}</div>' if sub else ""
    return (
        f'<div class="kpi-card">'
        f'<div class="kpi-val" style="color:{color}">{value}</div>'
        f'<div class="kpi-label">{label}</div>{sub_html}</div>'
    )


def _robust_to_numeric(series: pd.Series) -> pd.Series:
    from decimal import Decimal, InvalidOperation

    def _parse(v):
        if v is None:
            return np.nan
        if isinstance(v, bool):
            return np.nan
        if isinstance(v, (int, float)):
            if pd.isna(v):
                return np.nan
            return float(v)
        s = str(v).strip()
        if not s or s in ("-", "–", "—", "nan", "NaN", "None", "N/A",
                          "#N/A", "#REF!", "#DIV/0!", "#VALUE!", "#NAME?"):
            return np.nan
        s = re.sub(r"[₹$£€\s]", "", s)
        s = s.rstrip("%").strip()
        if not s:
            return np.nan
        s = s.replace(",", "")
        try:
            return float(Decimal(s))
        except (InvalidOperation, ValueError):
            return np.nan

    return series.apply(_parse)


# ─────────────────────────────────────────────────────────────────────────────
# OPERATIONS ENGINE
# ─────────────────────────────────────────────────────────────────────────────
OPERATIONS = [
    "Sum", "Mean (Avg)", "Median", "Min", "Max",
    "Count", "Std Deviation", "Variance", "Range (Max-Min)",
    "Top N Values", "Bottom N Values",
    "Cumulative Sum", "Rank (Desc)", "% of Total",
]


def run_op(df: pd.DataFrame, col: str, op: str,
           group_by: str = None, top_n: int = 10):
    if col not in df.columns:
        return None, f"Column '{col}' not found."
    numeric_series = _robust_to_numeric(df[col])
    valid = numeric_series.dropna()
    total = len(numeric_series)
    if valid.empty:
        return None, f"No numeric values found in column '{col}'."
    if group_by and group_by in df.columns:
        tmp = df[[group_by, col]].copy()
        tmp[col] = _robust_to_numeric(tmp[col])
        tmp = tmp.dropna(subset=[col])
        fn_map = {
            "Sum": "sum", "Mean (Avg)": "mean", "Median": "median",
            "Min": "min", "Max": "max", "Count": "count",
            "Std Deviation": "std", "Variance": "var",
            "Range (Max-Min)": lambda s: s.max() - s.min(),
        }
        fn = fn_map.get(op, "sum")
        if callable(fn):
            res = tmp.groupby(group_by)[col].apply(fn).reset_index()
        else:
            res = tmp.groupby(group_by)[col].agg(fn).reset_index()
        res = res.sort_values(col, ascending=False).reset_index(drop=True)
        res.index += 1
        valid_pct = f"{len(tmp) / total * 100:.1f}%" if total else "—"
        return res, f"{op} of '{col}' grouped by '{group_by}'", valid_pct, len(valid), total
    if op == "Sum":               v = valid.sum()
    elif op == "Mean (Avg)":      v = valid.mean()
    elif op == "Median":          v = valid.median()
    elif op == "Min":             v = valid.min()
    elif op == "Max":             v = valid.max()
    elif op == "Count":           v = float(len(valid))
    elif op == "Std Deviation":   v = valid.std(ddof=1)
    elif op == "Variance":        v = valid.var(ddof=1)
    elif op == "Range (Max-Min)": v = valid.max() - valid.min()
    elif op == "% of Total":
        grand = _robust_to_numeric(df[col]).dropna().sum()
        v = (valid.sum() / grand * 100) if grand else 0.0
    elif op in ("Top N Values", "Bottom N Values"):
        ctx = [c for c in ["_Location", "_Sheet"] if c in df.columns]
        cname = find_col(df, r"customer.*name|client.*name")
        if cname:
            ctx.append(cname)
        sub = df[ctx + [col]].copy()
        sub[col] = _robust_to_numeric(sub[col])
        sub = sub.dropna(subset=[col])
        if op == "Top N Values":
            sub = sub.nlargest(top_n, col)
        else:
            sub = sub.nsmallest(top_n, col)
        sub = sub.reset_index(drop=True)
        sub.index += 1
        valid_pct = f"{len(valid) / total * 100:.1f}%" if total else "—"
        return sub, f"{'Top' if 'Top' in op else 'Bottom'} {top_n} of '{col}'", valid_pct, len(valid), total
    elif op == "Cumulative Sum":
        sub = pd.DataFrame({
            "Row #": range(1, len(valid) + 1),
            col:     valid.values,
            "Cumulative Sum": valid.cumsum().values,
        })
        valid_pct = f"{len(valid) / total * 100:.1f}%" if total else "—"
        return sub, f"Cumulative Sum of '{col}'", valid_pct, len(valid), total
    elif op == "Rank (Desc)":
        ctx = [c for c in ["_Location"] if c in df.columns]
        cname = find_col(df, r"customer.*name|client.*name")
        if cname:
            ctx.append(cname)
        sub = df[ctx + [col]].copy()
        sub[col] = _robust_to_numeric(sub[col])
        sub = sub.dropna(subset=[col])
        sub["Rank"] = sub[col].rank(ascending=False, method="min").astype(int)
        sub = sub.sort_values("Rank").reset_index(drop=True)
        sub.index += 1
        valid_pct = f"{len(valid) / total * 100:.1f}%" if total else "—"
        return sub, f"Rank (Desc) by '{col}'", valid_pct, len(valid), total
    else:
        v = valid.sum()
    valid_pct = f"{len(valid) / total * 100:.1f}%" if total else "—"
    return float(v), f"{op} of '{col}'", valid_pct, len(valid), total


def _detect_unit(col_name: str) -> str:
    if not col_name:
        return ""
    c = col_name.lower()
    if "kva"  in c:                                           return "KVA"
    if any(k in c for k in ("kwhr", "kw hr", "kw-hr", "unit rate")):
        return "₹/kWh"
    if any(k in c for k in ("revenue", "mrc", "per unit rate", "charges", "billing")):
        return "₹"
    if any(k in c for k in ("kw", "kilowatt")):              return "kW"
    if any(k in c for k in ("sqft", "sq ft", "sq.ft", "subscription", "space", "area")):
        return "sq.ft"
    if "rack" in c:                                           return "racks"
    if any(k in c for k in ("seat", "sitting")):              return "seats"
    return ""


def _fmt_decimal(val: float, unit: str = "") -> str:
    if pd.isna(val):
        return "N/A"
    import math
    if val == 0:
        rounded = 0.0
    else:
        mag = math.floor(math.log10(abs(val)))
        rounded = round(val, max(0, 9 - mag))
    if rounded == int(rounded) and abs(rounded) < 1e12:
        disp_val = f"{int(rounded):,}"
    elif abs(rounded) >= 10_000:
        disp_val = f"{rounded:,.2f}"
    elif abs(rounded) >= 1:
        disp_val = f"{rounded:,.4f}".rstrip("0").rstrip(".")
    else:
        disp_val = f"{rounded:.6g}"
    return disp_val


# ─────────────────────────────────────────────────────────────────────────────
# CHART FACTORY
# ─────────────────────────────────────────────────────────────────────────────
CHART_TYPES = [
    "Bar Chart", "Grouped Bar", "Stacked Bar",
    "Line Chart", "Scatter Plot", "Area Chart",
    "Bubble Chart", "Heatmap (Correlation)", "Box Plot",
    "Violin Plot", "Funnel Chart", "Waterfall / Cumulative",
    "3-D Scatter", "Radar Chart", "Histogram",
]

CHART_NEEDS = {
    "Bar Chart":              {"x_cat", "y_num"},
    "Grouped Bar":            {"x_cat"},
    "Stacked Bar":            {"x_cat"},
    "Line Chart":             {"y_num"},
    "Scatter Plot":           {"x_num", "y_num", "color"},
    "Area Chart":             {"y_num"},
    "Bubble Chart":           {"x_num", "y_num", "size"},
    "Heatmap (Correlation)":  set(),
    "Box Plot":               {"x_cat", "y_num"},
    "Violin Plot":            {"x_cat", "y_num"},
    "Funnel Chart":           {"x_cat", "y_num"},
    "Waterfall / Cumulative": {"y_num"},
    "3-D Scatter":            {"x_num", "y_num", "z_num"},
    "Radar Chart":            set(),
    "Histogram":              {"y_num", "color"},
}


def _base_layout() -> dict:
    return dict(
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(15,22,40,0.85)",
        font=dict(color=TEXT, family="Segoe UI"),
        margin=dict(l=40, r=30, t=60, b=50),
    )


def make_chart(ct, df, x=None, y=None, color=None, size=None, z=None, title=""):
    lay = _base_layout()
    nc = num_cols(df)
    tc = txt_cols(df)
    if not x and tc: x = tc[0]
    if not y and nc: y = nc[0]
    try:
        if ct == "Bar Chart":
            if x and y:
                agg = df.groupby(x)[y].sum().reset_index().sort_values(y, ascending=False).head(30)
                fig = px.bar(agg, x=x, y=y, color=y, title=title, color_continuous_scale="Blues")
            else: fig = go.Figure()
        elif ct == "Grouped Bar":
            ys = [c for c in nc if c != x][:4]
            if x and ys:
                agg = df.groupby(x)[ys].sum().reset_index().head(20)
                fig = px.bar(agg, x=x, y=ys, barmode="group", title=title, color_discrete_sequence=px.colors.qualitative.Bold)
            else: fig = go.Figure()
        elif ct == "Stacked Bar":
            ys = [c for c in nc if c != x][:4]
            if x and ys:
                agg = df.groupby(x)[ys].sum().reset_index().head(20)
                fig = px.bar(agg, x=x, y=ys, barmode="stack", title=title, color_discrete_sequence=px.colors.qualitative.Pastel)
            else: fig = go.Figure()
        elif ct == "Line Chart":
            if y:
                sub = df[[c for c in [x, y] if c]].dropna().reset_index(drop=True)
                kw = dict(y=y, title=title, markers=True, color_discrete_sequence=[CYAN])
                if x: kw["x"] = x
                fig = px.line(sub, **kw)
            else: fig = go.Figure()
        elif ct == "Scatter Plot":
            if x and y:
                sub = df.dropna(subset=[c for c in [x, y] if c])
                kw = dict(x=x, y=y, title=title, opacity=0.7, color_discrete_sequence=[CYAN])
                if color and color in df.columns: kw["color"] = color; kw.pop("color_discrete_sequence")
                fig = px.scatter(sub, **kw)
            else: fig = go.Figure()
        elif ct == "Area Chart":
            if y:
                sub = df[[c for c in [x, y] if c]].dropna().reset_index(drop=True)
                kw = dict(y=y, title=title, color_discrete_sequence=[LBLUE])
                if x: kw["x"] = x
                fig = px.area(sub, **kw)
            else: fig = go.Figure()
        elif ct == "Heatmap (Correlation)":
            cols = nc[:14]
            if len(cols) >= 2:
                corr = df[cols].corr().round(2)
                fig = go.Figure(go.Heatmap(z=corr.values, x=corr.columns, y=corr.index, colorscale="RdBu", zmid=0, text=corr.values.round(2), texttemplate="%{text}"))
            else: fig = go.Figure()
        elif ct == "Box Plot":
            if y:
                kw = dict(y=y, title=title, color_discrete_sequence=px.colors.qualitative.Bold)
                if x: kw["x"] = x
                fig = px.box(df.dropna(subset=[y]), **kw)
            else: fig = go.Figure()
        elif ct == "Violin Plot":
            if y:
                kw = dict(y=y, title=title, box=True, points="outliers", color_discrete_sequence=[LBLUE])
                if x: kw["x"] = x
                fig = px.violin(df.dropna(subset=[y]), **kw)
            else: fig = go.Figure()
        elif ct == "Funnel Chart":
            if x and y:
                agg = df.groupby(x)[y].sum().reset_index().sort_values(y, ascending=False).head(20)
                fig = go.Figure(go.Funnel(y=agg[x].astype(str), x=agg[y], textinfo="value+percent total"))
            else: fig = go.Figure()
        elif ct == "Waterfall / Cumulative":
            if y:
                s = pd.to_numeric(df[y], errors="coerce").dropna().head(30)
                fig = go.Figure(go.Waterfall(x=list(range(len(s))), y=s.tolist(), measure=["relative"]*len(s), text=[f"{v:.1f}" for v in s], connector=dict(line=dict(color=BORD)), increasing=dict(marker_color=GREEN), decreasing=dict(marker_color=RED)))
            else: fig = go.Figure()
        elif ct == "3-D Scatter":
            if len(nc) >= 3:
                xc = x if x in nc else nc[0]; yc = y if y in nc else nc[1]; zc = z if z in nc else nc[2]
                sub = df[[xc, yc, zc]].dropna().head(500)
                fig = go.Figure(go.Scatter3d(x=sub[xc], y=sub[yc], z=sub[zc], mode="markers", marker=dict(size=5, color=sub[zc], colorscale="Blues", opacity=0.8, showscale=True)))
                fig.update_layout(scene=dict(xaxis_title=xc, yaxis_title=yc, zaxis_title=zc, bgcolor=DARK2))
            else: fig = go.Figure()
        elif ct == "Radar Chart":
            cols = nc[:8]
            if len(cols) >= 3:
                vals = df[cols].mean().tolist(); vals += [vals[0]]
                fig = go.Figure(go.Scatterpolar(r=vals, theta=cols+[cols[0]], fill="toself", line_color=CYAN, fillcolor="rgba(0,212,255,0.15)"))
                fig.update_layout(polar=dict(radialaxis=dict(visible=True, gridcolor=BORD), angularaxis=dict(gridcolor=BORD), bgcolor=DARK2))
            else: fig = go.Figure()
        elif ct == "Histogram":
            if y:
                kw = dict(x=y, nbins=30, title=title, opacity=0.85, color_discrete_sequence=[LBLUE])
                if color and color in df.columns: kw["color"] = color; kw.pop("color_discrete_sequence")
                fig = px.histogram(df.dropna(subset=[y]), **kw)
            else: fig = go.Figure()
        else: fig = go.Figure()
        fig.update_layout(title=title, **lay)
        return fig
    except Exception as exc:
        fig = go.Figure()
        fig.add_annotation(text=f"Chart error: {exc}", x=0.5, y=0.5, showarrow=False, font=dict(color=RED, size=14))
        fig.update_layout(**lay)
        return fig


# ─────────────────────────────────────────────────────────────────────────────
# LOAD DATA
# ─────────────────────────────────────────────────────────────────────────────
with st.spinner("Loading all Excel files…"):
    ALL = load_all()

# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(f"""
    <div style="text-align:center;padding:16px 0 20px">
      <div style="font-size:2.2rem">🏢</div>
      <div style="font-size:.95rem;font-weight:900;color:{WHITE};letter-spacing:.04em">
        SIFY DATA CENTRE</div>
      <div style="font-size:.7rem;color:{MUTED};margin-top:2px">
        Capacity Intelligence Platform</div>
    </div>""", unsafe_allow_html=True)

    all_locs = sorted(ALL.keys())
    if not all_locs:
        st.error("No Excel files found. Place your Excel files in the 'excel_files/' folder.")
        st.stop()

    sel_locs = st.multiselect("📍 Locations", all_locs, default=all_locs)
    all_sheet_opts = sorted({sn for loc in sel_locs for sn in ALL.get(loc, {})})
    sel_sheets = st.multiselect("📋 Sheets", all_sheet_opts, default=all_sheet_opts)

    st.markdown("---")
    n_loc = len(sel_locs)
    n_sh  = sum(len(ALL.get(l, {})) for l in sel_locs)
    st.markdown(
        f'<div style="font-size:.78rem;color:{MUTED}">Loaded '
        f'<b style="color:{CYAN}">{n_loc}</b> locations | '
        f'<b style="color:{CYAN}">{n_sh}</b> sheets</div>',
        unsafe_allow_html=True,
    )
    st.markdown("<br>", unsafe_allow_html=True)
    for loc in all_locs:
        n = len(ALL.get(loc, {}))
        active = "✓" if loc in sel_locs else "○"
        color = GREEN if loc in sel_locs else MUTED
        st.markdown(
            f'<div style="font-size:.76rem;color:{MUTED};padding:2px 0">'
            f'<span style="color:{color}">{active}</span> {loc} '
            f'<span style="color:{GREEN};font-weight:700">({n} sheets)</span></div>',
            unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# BUILD FILTERED DATA
# ─────────────────────────────────────────────────────────────────────────────
fdata = {loc: {sn: df for sn, df in sheets.items()
               if not sel_sheets or sn in sel_sheets}
         for loc, sheets in ALL.items() if loc in sel_locs}
fdata = {loc: sheets for loc, sheets in fdata.items() if sheets}

COMB = combined_df(fdata)

CUST_frames = []
for loc, sheets in fdata.items():
    for sn, df in sheets.items():
        tmp = df.copy()
        tmp.insert(0, "_Sheet", sn)
        tmp.insert(0, "_Location", loc)
        CUST_frames.append(tmp)

if CUST_frames:
    CUST = pd.concat(CUST_frames, ignore_index=True, sort=False).reset_index(drop=True)
else:
    CUST = pd.DataFrame()


# ─────────────────────────────────────────────────────────────────────────────
# HERO HEADER
# ─────────────────────────────────────────────────────────────────────────────
total_rec = len(CUST)
st.markdown(f"""
<div class="hero">
  <h1>🏢 Sify DC — Capacity Intelligence Platform</h1>
  <p>All {len(all_locs)} locations · {sum(len(s) for s in ALL.values())} sheets ·
     {total_rec:,} records loaded</p>
</div>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# TABS
# ─────────────────────────────────────────────────────────────────────────────
T = st.tabs(["📊 KPI Overview", "🗂 Data Explorer", "⚙️ Operations",
             "📈 Charts", "🔍 Smart Query", "🌐 Cross-Location"])


# ══════════════════════════════════════════════════════════════════════════════
# TAB 0 – KPI OVERVIEW (original cards + new Anvaya-style graphs)
# ══════════════════════════════════════════════════════════════════════════════
with T[0]:
    st.markdown('<div class="section-title">Key Performance Indicators — All Locations</div>',
                unsafe_allow_html=True)

    if CUST.empty:
        st.warning("No data loaded. Please check your Excel files.")
    else:
        nc = num_cols(CUST)
        tc = txt_cols(CUST)

        cust_c   = find_col(CUST, r"customer.*name|client.*name")
        caged_c  = find_col(CUST, r"\bcaged\b")
        own_c    = find_col(CUST, r"\brhs\b|\bshs\b|ownership")
        sub_mode_c = find_col(CUST, r"subscription.*mode\s*\(rack|space.*subscription.*mode")
        pw_sub_c = find_col(CUST, r"power.*subscription.*model|billing.*model.*power.*subscription")
        pw_use_m_c = find_col(CUST, r"power.*usage.*model|billing.*model.*power.*usage")

        space_sub_c  = find_col(CUST, r"space\s*\|\s*subscription$|^space.*subscription$")
        space_inuse_c = find_col(CUST, r"space.*in.*use")
        space_ytbg_c = find_col(CUST, r"yet.*to.*be.*given|yet.*billed")
        space_res_c  = find_col(CUST, r"reserved.*capacity")
        space_rate_c = find_col(CUST, r"per.*unit.*rate|per.*unit.*mrc")
        rack_c       = find_col(CUST, r"\brack\b")

        cap_c        = find_col(CUST, r"total.*capacity.*purchased|total.*capacity|capacity.*purchased")
        use_c        = find_col(CUST, r"capacity.*in.*use")
        cap_ytbg_c   = find_col(CUST, r"capacity.*to.*be.*given")
        cap_res_c    = find_col(CUST, r"reserved.*capacity")
        sub_kw_c     = find_col(CUST, r"subscribed.*capacity.*kw|capacity.*to.*be.*given.*kw")
        alloc_kw_c   = find_col(CUST, r"allocated.*capacity.*kw|\"allocated\".*kw|allocated.*kw")

        seat_sub_c   = find_col(CUST, r"seating.*space.*subscription|sitting.*space.*subscription|sitting.*space")
        seat_inuse_c = find_col(CUST, r"seating.*space.*in.*use|sitting.*space.*in.*use")

        rev_space_c  = find_col(CUST, r"space.*revenue.*including|space.*revenue")
        rev_addcap_c = find_col(CUST, r"additional.*capacity.*revenue|additional.*capacity.*charge")
        rev_pwuse_c  = find_col(CUST, r"power.*usage.*revenue")
        rev_seat_c   = find_col(CUST, r"seating.*space.*revenue|seating.*revenue")
        rev_other_c  = find_col(CUST, r"any.*other.*items|other.*items")
        rev_total_c  = find_col(CUST, r"total.*revenue")
        rev_freq_c   = find_col(CUST, r"billing.*frequency|frequency")
        rev_so_c     = find_col(CUST, r"sales.*order|so.*ref")

        con_start_c  = find_col(CUST, r"contract.*start|start.*date")
        con_term_c   = find_col(CUST, r"term.*contract|term.*year")
        con_expiry_c = find_col(CUST, r"current.*expiry|expiry.*date|expir")

        def _n(col):
            if col and col in CUST.columns:
                return _robust_to_numeric(CUST[col]).sum()
            return None

        def _avg(col):
            if col and col in CUST.columns:
                s = _robust_to_numeric(CUST[col]).dropna()
                return s.mean() if not s.empty else None
            return None

        def _cnt_val(col, val):
            if col and col in CUST.columns:
                return int((CUST[col].astype(str).str.upper().str.strip() == val.upper()).sum())
            return None

        # ── Top KPI Row ───────────────────────────────────────────────────
        k = st.columns(5)
        total_customers = CUST[cust_c].dropna().nunique() if cust_c else len(CUST)
        k[0].markdown(kpi_html(f"{total_customers:,}", "Unique Customers",
                               "across all locations", CYAN), unsafe_allow_html=True)
        if "_Location" in CUST.columns:
            k[1].markdown(kpi_html(f"{CUST['_Location'].nunique()}", "Active Locations",
                                   f"{sum(len(s) for s in fdata.values())} sheets", LBLUE),
                          unsafe_allow_html=True)
        k[2].markdown(kpi_html(f"{len(CUST):,}", "Total Records",
                               "All sheets combined", MUTED), unsafe_allow_html=True)
        if cap_c:
            k[3].markdown(kpi_html(fmt(_n(cap_c)), cap_c, "Power Capacity section", GREEN), unsafe_allow_html=True)
        if use_c:
            k[4].markdown(kpi_html(fmt(_n(use_c)), use_c, "Power Capacity section", AMBER), unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ── Billing Model ─────────────────────────────────────────────────
        st.markdown('<div class="section-title">Billing Model</div>', unsafe_allow_html=True)
        bm_cols = st.columns(4)
        if caged_c:
            cage_vals = CUST[caged_c].astype(str).str.upper().str.strip()
            n_caged   = (cage_vals == "CAGED").sum()
            n_uncaged = (cage_vals == "UNCAGED").sum()
            bm_cols[0].markdown(kpi_html(f"{n_caged}", caged_c, f"Uncaged: {n_uncaged}", CYAN), unsafe_allow_html=True)
        if own_c:
            own_vals = CUST[own_c].astype(str).str.strip().str.upper()
            n_sify     = int(own_vals.str.contains(r"SIFY", na=False).sum())
            n_customer = int(own_vals.str.contains(r"CUSTOMER|CUST(?!OM)", na=False).sum())
            if n_sify > 0 or n_customer > 0:
                bm_cols[1].markdown(kpi_html(f"Sify: {n_sify}", "Space | Ownership", f"Customer: {n_customer}", LBLUE), unsafe_allow_html=True)
            else:
                rhs_c_cnt = _cnt_val(own_c, "RHS")
                shs_c_cnt = _cnt_val(own_c, "SHS")
                if rhs_c_cnt is not None or shs_c_cnt is not None:
                    bm_cols[1].markdown(kpi_html(f"RHS: {rhs_c_cnt or 0}", "Space | Ownership", f"SHS: {shs_c_cnt or 0}", LBLUE), unsafe_allow_html=True)
        if pw_sub_c:
            rated = _cnt_val(pw_sub_c, "RATED")
            subsc = _cnt_val(pw_sub_c, "SUBSCRIBED")
            bm_cols[2].markdown(kpi_html(f"{rated or 0}", pw_sub_c, f"Subscribed: {subsc or 0}", AMBER), unsafe_allow_html=True)
        if pw_use_m_c:
            bundled = _cnt_val(pw_use_m_c, "BUNDLED")
            metered = _cnt_val(pw_use_m_c, "METERED")
            bm_cols[3].markdown(kpi_html(f"{bundled or 0}", pw_use_m_c, f"Metered: {metered or 0}", GREEN), unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ── Space ─────────────────────────────────────────────────────────
        st.markdown('<div class="section-title">Space</div>', unsafe_allow_html=True)
        sp_cols = st.columns(5)
        if sub_mode_c:
            sub_vals_upper = CUST[sub_mode_c].astype(str).str.strip().str.upper()
            rack_m = _cnt_val(sub_mode_c, "RACK"); u_m = 0; sqft_m = 0
            sp_cols[0].markdown(kpi_html(f"{rack_m or 0}", sub_mode_c, f"Rack subscriptions", CYAN), unsafe_allow_html=True)
        if space_sub_c:
            v = _n(space_sub_c)
            if v is not None: sp_cols[1].markdown(kpi_html(fmt(v), space_sub_c, space_sub_c[:25], GREEN), unsafe_allow_html=True)
        if space_inuse_c:
            v = _n(space_inuse_c)
            if v is not None: sp_cols[2].markdown(kpi_html(fmt(v), space_inuse_c, space_inuse_c[:25], AMBER), unsafe_allow_html=True)
        if space_ytbg_c:
            v = _n(space_ytbg_c)
            if v is not None: sp_cols[3].markdown(kpi_html(fmt(v), space_ytbg_c, space_ytbg_c[:25], RED), unsafe_allow_html=True)
        if space_rate_c:
            v = _avg(space_rate_c)
            if v is not None: sp_cols[4].markdown(kpi_html(fmt(v), space_rate_c, space_rate_c[:25], LBLUE), unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ── Power Capacity ────────────────────────────────────────────────
        st.markdown('<div class="section-title">Power Capacity</div>', unsafe_allow_html=True)
        pc_cols = st.columns(5)
        if cap_c: pc_cols[0].markdown(kpi_html(fmt(_n(cap_c)), cap_c, cap_c[:25], GREEN), unsafe_allow_html=True)
        if use_c: pc_cols[1].markdown(kpi_html(fmt(_n(use_c)), use_c, use_c[:25], AMBER), unsafe_allow_html=True)
        if cap_ytbg_c:
            v = _n(cap_ytbg_c)
            if v is not None: pc_cols[2].markdown(kpi_html(fmt(v), cap_ytbg_c, cap_ytbg_c[:25], RED), unsafe_allow_html=True)
        if sub_kw_c:
            v = _n(sub_kw_c)
            if v is not None: pc_cols[3].markdown(kpi_html(fmt(v), sub_kw_c, sub_kw_c[:25], LBLUE), unsafe_allow_html=True)
        if alloc_kw_c:
            v = _n(alloc_kw_c)
            if v is not None: pc_cols[4].markdown(kpi_html(fmt(v), alloc_kw_c, alloc_kw_c[:25], CYAN), unsafe_allow_html=True)
        elif cap_c and use_c:
            t_cap = _n(cap_c) or 0; t_use = _n(use_c) or 0
            util = (t_use / t_cap * 100) if t_cap > 0 else 0
            pc_cols[4].markdown(kpi_html(f"{util:.1f}%", "Utilisation Rate", "Capacity In Use / Purchased", AMBER), unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ── Revenue (Monthly) ─────────────────────────────────────────────
        st.markdown('<div class="section-title">Revenue (Monthly)</div>', unsafe_allow_html=True)
        rv_cols = st.columns(5)
        rv_items = [(rev_space_c, CYAN), (rev_addcap_c, LBLUE), (rev_pwuse_c, GREEN), (rev_seat_c, AMBER), (rev_other_c, MUTED)]
        filled = 0
        for col, color in rv_items:
            if col and filled < 5:
                v = _n(col)
                if v is not None:
                    rv_cols[filled].markdown(kpi_html(fmt(v), col, col[:25], color), unsafe_allow_html=True)
                    filled += 1

        st.markdown("<br>", unsafe_allow_html=True)
        rv2_cols = st.columns(4)
        if rev_total_c:
            v = _n(rev_total_c)
            if v is not None: rv2_cols[0].markdown(kpi_html(fmt(v), rev_total_c, rev_total_c[:25], GREEN), unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ── Contract Information ──────────────────────────────────────────
        st.markdown('<div class="section-title">Contract Information</div>', unsafe_allow_html=True)
        ci_cols = st.columns(4)
        ci_i = 0
        if con_start_c:
            non_null = CUST[con_start_c].dropna()
            ci_cols[ci_i].markdown(kpi_html(f"{len(non_null):,}", con_start_c, con_start_c[:25], CYAN), unsafe_allow_html=True)
            ci_i += 1
        if con_term_c:
            v = _avg(con_term_c)
            if v is not None:
                ci_cols[ci_i].markdown(kpi_html(f"{v:.1f} yr", con_term_c, con_term_c[:25], GREEN), unsafe_allow_html=True)
                ci_i += 1
        if con_expiry_c:
            non_null = CUST[con_expiry_c].dropna()
            ci_cols[ci_i].markdown(kpi_html(f"{len(non_null):,}", con_expiry_c, con_expiry_c[:25], AMBER), unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ── Per-Location Summary Table ────────────────────────────────────
        st.markdown('<div class="section-title">Per-Location Summary</div>', unsafe_allow_html=True)
        if "_Location" in CUST.columns:
            agg_cols = [c for c in [cap_c, use_c, rev_total_c, rev_space_c, rev_pwuse_c] if c]
            if agg_cols:
                loc_agg = CUST.groupby("_Location")[agg_cols].apply(
                    lambda g: g.apply(pd.to_numeric, errors="coerce").sum()
                ).reset_index()
                loc_agg.columns = ["Location"] + agg_cols
                if cust_c:
                    loc_agg["Customer Count"] = (CUST.groupby("_Location")[cust_c].apply(lambda g: g.dropna().nunique()).values)
                st.dataframe(loc_agg.round(2), use_container_width=True)

        # ── Utilisation Gauges ────────────────────────────────────────────
        if cap_c and use_c:
            st.markdown('<div class="section-title">Utilisation Gauges</div>', unsafe_allow_html=True)
            g1, g2 = st.columns(2)
            t_cap = _n(cap_c) or 0; t_use = _n(use_c) or 0
            util_pct = min((t_use / t_cap * 100) if t_cap > 0 else 0, 100)
            t_space_sub = _n(space_sub_c) or 0; t_space_use = _n(space_inuse_c) or 0
            rack_pct = min((t_space_use / t_space_sub * 100) if t_space_sub > 0 else util_pct, 100)

            def _gauge(val, label, bar_color):
                fig = go.Figure(go.Indicator(
                    mode="gauge+number", value=min(float(val), 100),
                    title={"text": label, "font": {"color": TEXT, "size": 14}},
                    gauge={"axis": {"range": [0, 100], "tickcolor": TEXT}, "bar": {"color": bar_color}, "bgcolor": DARK2,
                        "steps": [{"range": [0, 50], "color": "#1a2a1a"}, {"range": [50, 80], "color": "#2a2a1a"}, {"range": [80, 100], "color": "#2a1a1a"}],
                        "threshold": {"line": {"color": RED, "width": 3}, "value": 80}},
                    number={"suffix": "%", "font": {"color": bar_color}}))
                fig.update_layout(**_base_layout(), height=270)
                return fig

            g1.plotly_chart(_gauge(util_pct, "Capacity Utilisation (%)", LBLUE), use_container_width=True)
            g2.plotly_chart(_gauge(rack_pct, "Space/Rack Utilisation (%)", GREEN), use_container_width=True)

        # ── Capacity vs Usage by Location ─────────────────────────────────
        if cap_c and use_c and "_Location" in CUST.columns:
            st.markdown('<div class="section-title">Capacity vs Usage by Location</div>', unsafe_allow_html=True)
            la = CUST.groupby("_Location").agg(
                Capacity_Purchased=(cap_c, lambda x: pd.to_numeric(x, errors="coerce").sum()),
                Capacity_in_Use=(use_c, lambda x: pd.to_numeric(x, errors="coerce").sum()),
            ).reset_index()
            fig_la = px.bar(la, x="_Location", y=["Capacity_Purchased", "Capacity_in_Use"], barmode="group",
                labels={"_Location": "Location", "value": "Units"},
                color_discrete_map={"Capacity_Purchased": LBLUE, "Capacity_in_Use": GREEN})
            fig_la.update_layout(**_base_layout(), height=360)
            st.plotly_chart(fig_la, use_container_width=True)

        # ── Space & Revenue Split Pie Charts ──────────────────────────────
        st.markdown('<div class="section-title">Space &amp; Revenue Split</div>', unsafe_allow_html=True)
        pie_cols = st.columns(3)
        if caged_c:
            cv = CUST[caged_c].astype(str).str.upper().str.strip()
            pie_d = cv.value_counts().reset_index(); pie_d.columns = ["Status", "Count"]
            if not pie_d.empty:
                fig_p1 = px.pie(pie_d, names="Status", values="Count", title="Caged vs Uncaged", color_discrete_sequence=[CYAN, LBLUE, GREEN, AMBER])
                fig_p1.update_layout(**_base_layout(), height=300)
                pie_cols[0].plotly_chart(fig_p1, use_container_width=True)
        if pw_sub_c:
            pie_d2 = CUST[pw_sub_c].dropna().value_counts().reset_index(); pie_d2.columns = ["Model", "Count"]
            if not pie_d2.empty:
                fig_p2 = px.pie(pie_d2, names="Model", values="Count", title="Power Subscription Model", color_discrete_sequence=[LBLUE, GREEN, AMBER, RED])
                fig_p2.update_layout(**_base_layout(), height=300)
                pie_cols[1].plotly_chart(fig_p2, use_container_width=True)
        if pw_use_m_c:
            pie_d3 = CUST[pw_use_m_c].dropna().value_counts().reset_index(); pie_d3.columns = ["Model", "Count"]
            if not pie_d3.empty:
                fig_p3 = px.pie(pie_d3, names="Model", values="Count", title="Power Usage Model", color_discrete_sequence=[GREEN, AMBER, CYAN, RED])
                fig_p3.update_layout(**_base_layout(), height=300)
                pie_cols[2].plotly_chart(fig_p3, use_container_width=True)

        # ══════════════════════════════════════════════════════════════════
        # NEW: ANVAYA-STYLE GRAPHS — Power / Rack Space / White Space
        # ══════════════════════════════════════════════════════════════════

        def _extract_fac_kpi(all_data_dict):
            fac = {}
            for loc, sheets in all_data_dict.items():
                info = {"location": loc, "kw_sold": 0, "kw_consumed": 0,
                        "space_total": 0, "space_used": 0,
                        "rack_sold": 0, "rack_operational": 0}
                for sn, df in sheets.items():
                    cap_c2 = use_c2 = ssub_c2 = suse_c2 = rsub_c2 = ruse_c2 = None
                    for c in df.columns:
                        cl = c.lower()
                        if re.search(r"total.*capacity.*purchased|total.*capacity.*kw", cl) and not cap_c2: cap_c2 = c
                        elif re.search(r"capacity.*in.*use|capacity.*in.*use.*kw", cl) and not use_c2: use_c2 = c
                        elif re.search(r"space.*\|.*subscription$|^space.*subscription$", cl) and not ssub_c2: ssub_c2 = c
                        elif re.search(r"space.*\|.*in.*use$|^space.*in.*use$", cl) and not suse_c2: suse_c2 = c
                        elif re.search(r"subscription.*no.*rack|rack.*subscription|subscription\(no", cl) and not rsub_c2: rsub_c2 = c
                        elif re.search(r"in.*use.*no.*rack|rack.*in.*use|in.*use\(no", cl) and not ruse_c2: ruse_c2 = c
                    if cap_c2 and cap_c2 in df.columns: info["kw_sold"] += _robust_to_numeric(df[cap_c2]).dropna().sum()
                    if use_c2 and use_c2 in df.columns: info["kw_consumed"] += _robust_to_numeric(df[use_c2]).dropna().sum()
                    if ssub_c2 and ssub_c2 in df.columns: info["space_total"] += _robust_to_numeric(df[ssub_c2]).dropna().sum()
                    if suse_c2 and suse_c2 in df.columns: info["space_used"] += _robust_to_numeric(df[suse_c2]).dropna().sum()
                    if rsub_c2 and rsub_c2 in df.columns: info["rack_sold"] += _robust_to_numeric(df[rsub_c2]).dropna().sum()
                    if ruse_c2 and ruse_c2 in df.columns: info["rack_operational"] += _robust_to_numeric(df[ruse_c2]).dropna().sum()
                info["kw_available"] = max(0, info["kw_sold"] - info["kw_consumed"])
                info["space_available"] = max(0, info["space_total"] - info["space_used"])
                info["rack_available"] = max(0, info["rack_sold"] - info["rack_operational"])
                fac[loc] = info
            return fac

        _fac_data = _extract_fac_kpi(fdata)

        # ── RACK POWER USAGE (kW) ────────────────────────────────────────
        st.markdown('<div class="section-title">Rack Power Usage (kW) &mdash; Subscribed vs In Use by Location</div>', unsafe_allow_html=True)
        _rpu_rows = []
        for _fl, _fi in sorted(_fac_data.items()):
            if _fi["kw_sold"] > 0 or _fi["kw_consumed"] > 0:
                _rpu_rows.append({"Location": _fl, "Subscribed (kW)": round(_fi["kw_sold"], 2),
                    "In Use (kW)": round(_fi["kw_consumed"], 2), "Available (kW)": round(_fi["kw_available"], 2)})
        if _rpu_rows:
            _rpu_df = pd.DataFrame(_rpu_rows)
            _rpu_ts = _rpu_df["Subscribed (kW)"].sum(); _rpu_tu = _rpu_df["In Use (kW)"].sum()
            _rpu_ta = _rpu_df["Available (kW)"].sum()
            _rpu_pct = (_rpu_tu / _rpu_ts * 100) if _rpu_ts > 0 else 0
            _rpk1, _rpk2, _rpk3, _rpk4 = st.columns(4)
            _rpk1.markdown(kpi_html(f"{_rpu_ts:,.0f} kW", "Total Subscribed", "All Locations", CYAN), unsafe_allow_html=True)
            _rpk2.markdown(kpi_html(f"{_rpu_tu:,.0f} kW", "Total In Use", f"{_rpu_pct:.1f}% Utilised", AMBER), unsafe_allow_html=True)
            _rpk3.markdown(kpi_html(f"{_rpu_ta:,.0f} kW", "Total Available", f"{100-_rpu_pct:.1f}% Free", GREEN), unsafe_allow_html=True)
            _rpk4.markdown(kpi_html(f"{len(_rpu_rows)}", "Locations", "With power data", LBLUE), unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)
            fig_rpu = go.Figure()
            fig_rpu.add_trace(go.Bar(x=_rpu_df["Location"], y=_rpu_df["Subscribed (kW)"], name="Subscribed", marker_color="#2196F3",
                text=_rpu_df["Subscribed (kW)"].apply(lambda v: f"{v:,.0f}"), textposition="outside", textfont=dict(size=9)))
            fig_rpu.add_trace(go.Bar(x=_rpu_df["Location"], y=_rpu_df["In Use (kW)"], name="In Use", marker_color="#FF5722",
                text=_rpu_df["In Use (kW)"].apply(lambda v: f"{v:,.0f}"), textposition="outside", textfont=dict(size=9)))
            fig_rpu.add_trace(go.Bar(x=_rpu_df["Location"], y=_rpu_df["Available (kW)"], name="Available", marker_color="#4CAF50",
                text=_rpu_df["Available (kW)"].apply(lambda v: f"{v:,.0f}"), textposition="outside", textfont=dict(size=9)))
            fig_rpu.update_layout(barmode="group", **_base_layout(), height=420, title="Power Capacity by Location (kW)",
                xaxis_title="DC Location", yaxis_title="kW",
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5), xaxis_tickangle=-25)
            st.plotly_chart(fig_rpu, use_container_width=True)
            st.dataframe(_rpu_df.set_index("Location"), use_container_width=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ── WHITE SPACE + RACK SPACE ─────────────────────────────────────
        st.markdown('<div class="section-title">White Space &amp; Rack Space &mdash; by Location</div>', unsafe_allow_html=True)
        _rs_rows = []
        for _fl, _fi in sorted(_fac_data.items()):
            if _fi["space_total"] > 0 or _fi["space_used"] > 0 or _fi["rack_sold"] > 0:
                _rs_rows.append({"Location": _fl, "Space Subscribed": round(_fi["space_total"], 2),
                    "Space In Use": round(_fi["space_used"], 2), "Space Available": round(_fi["space_available"], 2),
                    "Racks Subscribed": round(_fi["rack_sold"], 2), "Racks In Use": round(_fi["rack_operational"], 2),
                    "Racks Available": round(_fi["rack_available"], 2)})
        if _rs_rows:
            _rs_df = pd.DataFrame(_rs_rows)
            _rsc1, _rsc2 = st.columns(2)
            with _rsc1:
                _ts_sub = _rs_df["Space Subscribed"].sum(); _ts_use = _rs_df["Space In Use"].sum()
                _ts_avl = _rs_df["Space Available"].sum()
                _ts_pct = (_ts_use / _ts_sub * 100) if _ts_sub > 0 else 0
                st.markdown(f'<div style="text-align:center;font-size:.78rem;color:{MUTED};margin-bottom:8px">'
                    f'Total: <b style="color:{CYAN}">{_ts_sub:,.0f}</b> | In Use: <b style="color:{AMBER}">{_ts_use:,.0f}</b> '
                    f'| Avail: <b style="color:{GREEN}">{_ts_avl:,.0f}</b> ({100-_ts_pct:.1f}% free)</div>', unsafe_allow_html=True)
                fig_ws = go.Figure()
                fig_ws.add_trace(go.Bar(x=_rs_df["Location"], y=_rs_df["Space Subscribed"], name="Subscribed", marker_color="#2196F3"))
                fig_ws.add_trace(go.Bar(x=_rs_df["Location"], y=_rs_df["Space In Use"], name="In Use", marker_color="#FF5722"))
                fig_ws.add_trace(go.Bar(x=_rs_df["Location"], y=_rs_df["Space Available"], name="Available", marker_color="#4CAF50"))
                fig_ws.update_layout(barmode="group", **_base_layout(), height=380, title="White Space (Sq.Ft / Racks)",
                    xaxis_tickangle=-30, legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5))
                st.plotly_chart(fig_ws, use_container_width=True)
            with _rsc2:
                _rk_df = _rs_df[(_rs_df["Racks Subscribed"] > 0) | (_rs_df["Racks In Use"] > 0)]
                _trk_sub = _rk_df["Racks Subscribed"].sum() if not _rk_df.empty else 0
                _trk_use = _rk_df["Racks In Use"].sum() if not _rk_df.empty else 0
                _trk_avl = _rk_df["Racks Available"].sum() if not _rk_df.empty else 0
                _trk_pct = (_trk_use / _trk_sub * 100) if _trk_sub > 0 else 0
                st.markdown(f'<div style="text-align:center;font-size:.78rem;color:{MUTED};margin-bottom:8px">'
                    f'Total: <b style="color:{CYAN}">{_trk_sub:,.0f}</b> | In Use: <b style="color:{AMBER}">{_trk_use:,.0f}</b> '
                    f'| Avail: <b style="color:{GREEN}">{_trk_avl:,.0f}</b> ({100-_trk_pct:.1f}% free)</div>', unsafe_allow_html=True)
                if not _rk_df.empty:
                    fig_rk = go.Figure()
                    fig_rk.add_trace(go.Bar(x=_rk_df["Location"], y=_rk_df["Racks Subscribed"], name="Subscribed", marker_color="#2196F3"))
                    fig_rk.add_trace(go.Bar(x=_rk_df["Location"], y=_rk_df["Racks In Use"], name="In Use", marker_color="#FF5722"))
                    fig_rk.add_trace(go.Bar(x=_rk_df["Location"], y=_rk_df["Racks Available"], name="Available", marker_color="#4CAF50"))
                    fig_rk.update_layout(barmode="group", **_base_layout(), height=380, title="Rack Space (Count)",
                        xaxis_tickangle=-30, legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5))
                    st.plotly_chart(fig_rk, use_container_width=True)
            st.dataframe(_rs_df.set_index("Location"), use_container_width=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ── PER-SITE SUMMARY CARDS (Anvaya-style) ────────────────────────
        st.markdown('<div class="section-title">Per-Site Summary Cards</div>', unsafe_allow_html=True)
        _sc_list = []
        for _fl, _fi in sorted(_fac_data.items()):
            kw_s = _fi["kw_sold"]; kw_u = _fi["kw_consumed"]
            sp_s = _fi["space_total"]; sp_u = _fi["space_used"]
            rk_s = _fi["rack_sold"]; rk_u = _fi["rack_operational"]
            pwr_a = round((kw_s - kw_u) / kw_s * 100, 1) if kw_s > 0 else 0
            spc_a = round((sp_s - sp_u) / sp_s * 100, 1) if sp_s > 0 else 0
            rck_a = round((rk_s - rk_u) / rk_s * 100, 1) if rk_s > 0 else 0
            ups_l = round(kw_u / kw_s * 100, 2) if kw_s > 0 else 0
            _sc_list.append(dict(loc=_fl, kw_s=kw_s, kw_u=kw_u, pwr_a=pwr_a,
                ups_l=ups_l, rck_a=rck_a, spc_a=spc_a))
        for _ci in range(0, len(_sc_list), 3):
            _batch = _sc_list[_ci:_ci+3]
            _ccols = st.columns(3)
            for _j, _cd in enumerate(_batch):
                with _ccols[_j]:
                    _pc = GREEN if _cd["pwr_a"] > 20 else (AMBER if _cd["pwr_a"] > 5 else RED)
                    _rc = GREEN if _cd["rck_a"] > 20 else (AMBER if _cd["rck_a"] > 5 else RED)
                    _scc = GREEN if _cd["spc_a"] > 20 else (AMBER if _cd["spc_a"] > 5 else RED)
                    _card_html = (
                        f'<div style="background:{CARD};border:1px solid {BORD};border-radius:14px;'
                        f'padding:18px 20px;margin:6px 0;box-shadow:0 4px 16px rgba(0,0,0,.3)">'
                        f'<div style="font-size:1rem;font-weight:900;color:{WHITE};margin-bottom:12px;'
                        f'border-bottom:1px solid {BORD};padding-bottom:8px">'
                        f'&#127970; {_cd["loc"]}</div>'
                        f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;font-size:.76rem">'
                        f'<div><span style="color:{MUTED}">Power Subscribed:</span><br>'
                        f'<b style="color:{CYAN}">{_cd["kw_s"]:,.0f} kW</b></div>'
                        f'<div><span style="color:{MUTED}">Power In Use:</span><br>'
                        f'<b style="color:{AMBER}">{_cd["kw_u"]:,.0f} kW</b></div>'
                        f'<div><span style="color:{MUTED}">Ups Load:</span><br>'
                        f'<b style="color:{AMBER}">{_cd["ups_l"]:.1f}%</b></div>'
                        f'<div><span style="color:{MUTED}">Power Avl:</span><br>'
                        f'<b style="color:{_pc}">{_cd["pwr_a"]:.0f}%</b></div>'
                        f'<div><span style="color:{MUTED}">Rack Avl:</span><br>'
                        f'<b style="color:{_rc}">{_cd["rck_a"]:.0f}%</b></div>'
                        f'<div><span style="color:{MUTED}">Space Avl:</span><br>'
                        f'<b style="color:{_scc}">{_cd["spc_a"]:.0f}%</b></div>'
                        f'</div></div>')
                    st.markdown(_card_html, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ── REVENUE BY LOCATION — Stacked Horizontal Bar ─────────────────
        if rev_total_c or rev_space_c or rev_pwuse_c:
            st.markdown('<div class="section-title">Revenue Breakdown by Location (Monthly MRC)</div>', unsafe_allow_html=True)
            _rv_avail = [(rev_space_c, "Space Revenue", "#2196F3"), (rev_addcap_c, "Add Cap Revenue", "#FF9800"),
                         (rev_pwuse_c, "Power Revenue", "#4CAF50"), (rev_seat_c, "Seating Revenue", "#9C27B0"),
                         (rev_other_c, "Other", "#607D8B")]
            _rv_avail = [(c,n,cl) for c,n,cl in _rv_avail if c]
            if _rv_avail and "_Location" in CUST.columns:
                _rv_rows = []
                for _rl in sorted(CUST["_Location"].unique()):
                    _rl_df = CUST[CUST["_Location"] == _rl]
                    _rw = {"Location": _rl}
                    for _rc, _rn, _rcl in _rv_avail:
                        _rw[_rn] = round(_robust_to_numeric(_rl_df[_rc]).dropna().sum(), 2) if _rc in _rl_df.columns else 0
                    _rv_rows.append(_rw)
                _rv_df = pd.DataFrame(_rv_rows)
                fig_rv = go.Figure()
                for _rc, _rn, _rcl in _rv_avail:
                    if _rn in _rv_df.columns:
                        fig_rv.add_trace(go.Bar(y=_rv_df["Location"], x=_rv_df[_rn], name=_rn, orientation="h", marker_color=_rcl))
                fig_rv.update_layout(barmode="stack", **_base_layout(), height=max(350, len(_rv_rows)*45),
                    title="Revenue Components by Location", xaxis_title="Amount (Monthly MRC)",
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5))
                st.plotly_chart(fig_rv, use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 – DATA EXPLORER
# ══════════════════════════════════════════════════════════════════════════════
with T[1]:
    st.markdown('<div class="section-title">Data Explorer</div>', unsafe_allow_html=True)
    if CUST.empty:
        st.warning("No data loaded.")
    else:
        de1, de2, de3 = st.columns(3)
        with de1: de_loc = st.selectbox("Location", ["All"] + sorted(fdata.keys()), key="de_loc")
        with de2:
            sheet_opts = sorted(fdata.get(de_loc, {}).keys()) if de_loc != "All" else sorted({sn for s in fdata.values() for sn in s})
            de_sh = st.selectbox("Sheet", ["All"] + sheet_opts, key="de_sh")
        with de3: de_search = st.text_input("Search", key="de_search", placeholder="type to filter…")
        view_df = CUST.copy()
        if de_loc != "All" and "_Location" in view_df.columns: view_df = view_df[view_df["_Location"] == de_loc]
        if de_sh != "All" and "_Sheet" in view_df.columns: view_df = view_df[view_df["_Sheet"] == de_sh]
        if de_search.strip():
            mask = view_df.apply(lambda r: r.astype(str).str.lower().str.contains(de_search.lower(), na=False).any(), axis=1)
            view_df = view_df[mask]
        st.markdown(f'<span class="badge">{len(view_df):,} rows</span> <span class="badge" style="background:{DARK2}">{len(view_df.columns)} cols</span>', unsafe_allow_html=True)
        st.dataframe(view_df.head(1000), use_container_width=True, height=480)
        st.download_button("Download CSV", view_df.to_csv(index=False).encode(), "sify_data.csv", "text/csv")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 – OPERATIONS
# ══════════════════════════════════════════════════════════════════════════════
with T[2]:
    st.markdown('<div class="section-title">Operations Engine</div>', unsafe_allow_html=True)
    if CUST.empty:
        st.warning("No data loaded.")
    else:
        op1, op2 = st.columns(2)
        with op1: op_loc = st.selectbox("Location", ["All"] + sorted(fdata.keys()), key="op_loc")
        with op2:
            op_sh_opts = sorted(fdata.get(op_loc, {}).keys()) if op_loc != "All" else sorted({sn for s in fdata.values() for sn in s})
            op_sh = st.selectbox("Sheet", ["All"] + op_sh_opts, key="op_sh")
        op_df = CUST.copy()
        if op_loc != "All" and "_Location" in op_df.columns: op_df = op_df[op_df["_Location"].str.contains(op_loc, case=False, na=False)]
        if op_sh != "All" and "_Sheet" in op_df.columns: op_df = op_df[op_df["_Sheet"] == op_sh]
        st.caption(f"**{len(op_df):,}** rows available")
        all_cols = [c for c in op_df.columns if not c.startswith("_")]
        nc_op = num_cols(op_df)
        grp_candidates = [c for c in ["_Location", "_Sheet"] if c in op_df.columns] + [c for c in txt_cols(op_df)]
        op3, op4, op5 = st.columns([2, 2, 2])
        with op3: op_col = st.selectbox("Column", all_cols, key="op_col")
        with op4: op_op = st.selectbox("Operation", OPERATIONS, key="op_op")
        with op5: op_grp = st.selectbox("Group By", ["None"] + [c for c in grp_candidates if c != op_col], key="op_grp")
        op_n = st.number_input("N (for Top/Bottom)", min_value=1, max_value=500, value=10, step=1, key="op_n")
        if st.button("Run Operation", key="op_run", use_container_width=True):
            if op_col and op_col in op_df.columns:
                grp = op_grp if op_grp != "None" else None
                out = run_op(op_df, op_col, op_op, grp, int(op_n))
                if out[0] is None:
                    st.error(out[1])
                else:
                    result, desc, valid_pct, valid_count, total_count = out
                    st.markdown(f'<div class="section-title">{desc}</div>', unsafe_allow_html=True)
                    if isinstance(result, (int, float)):
                        unit = _detect_unit(op_col)
                        st.markdown(f'<div class="result-box"><div class="result-big">{_fmt_decimal(result)} {unit}</div>'
                            f'<div style="font-size:.8rem;color:{MUTED};margin-top:6px">{valid_count:,} / {total_count:,} rows ({valid_pct})</div></div>', unsafe_allow_html=True)
                    elif isinstance(result, pd.DataFrame):
                        st.dataframe(result, use_container_width=True)
                        if grp and op_col in result.columns and grp in result.columns:
                            fig_op = px.bar(result.head(30), x=grp, y=op_col, color=op_col, color_continuous_scale="Blues", title=desc)
                            fig_op.update_layout(**_base_layout(), height=420)
                            st.plotly_chart(fig_op, use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 – CHARTS
# ══════════════════════════════════════════════════════════════════════════════
with T[3]:
    st.markdown('<div class="section-title">Chart Studio</div>', unsafe_allow_html=True)
    if CUST.empty:
        st.warning("No data loaded.")
    else:
        ch1, ch2, ch3 = st.columns(3)
        with ch1: ch_loc = st.selectbox("Location", ["All"] + sorted(fdata.keys()), key="ch_loc")
        with ch2:
            ch_sh_opts = sorted(fdata.get(ch_loc, {}).keys()) if ch_loc != "All" else sorted({sn for s in fdata.values() for sn in s})
            ch_sh = st.selectbox("Sheet", ["All"] + ch_sh_opts, key="ch_sh")
        with ch3: ch_type = st.selectbox("Chart Type", CHART_TYPES, key="ch_type")
        ch_df = CUST.copy()
        if ch_loc != "All" and "_Location" in ch_df.columns: ch_df = ch_df[ch_df["_Location"] == ch_loc]
        if ch_sh != "All" and "_Sheet" in ch_df.columns: ch_df = ch_df[ch_df["_Sheet"] == ch_sh]
        nc_ch = num_cols(ch_df); tc_ch = txt_cols(ch_df)
        needs = CHART_NEEDS.get(ch_type, set())
        ca, cb, cc, cd = st.columns(4)
        x_val = ca.selectbox("X-axis", ["—"] + tc_ch + nc_ch, key="ch_x") if "x_cat" in needs or "x_num" in needs else None
        y_val = cb.selectbox("Y-axis", ["—"] + nc_ch, key="ch_y") if "y_num" in needs else None
        col_val = cc.selectbox("Color", ["—"] + tc_ch + nc_ch, key="ch_col") if "color" in needs else None
        sz_val = cd.selectbox("Size", ["—"] + nc_ch, key="ch_sz") if "size" in needs else None
        z_val = ca.selectbox("Z-axis", ["—"] + nc_ch, key="ch_z") if "z_num" in needs else None
        if st.button("Generate Chart", key="ch_run"):
            kw = dict(x=x_val if x_val and x_val != "—" else None, y=y_val if y_val and y_val != "—" else None,
                color=col_val if col_val and col_val != "—" else None, size=sz_val if sz_val and sz_val != "—" else None,
                z=z_val if z_val and z_val != "—" else None, title=f"{ch_type} — {ch_loc} / {ch_sh}")
            fig = make_chart(ch_type, ch_df, **kw)
            fig.update_layout(height=500)
            st.plotly_chart(fig, use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 – SMART QUERY (Customer Search + Column Operations)
# ══════════════════════════════════════════════════════════════════════════════
with T[4]:
    st.markdown('<div class="section-title">Smart Query — Customer Search &amp; Column Operations</div>', unsafe_allow_html=True)
    _sq_mode = st.radio("Mode", ["Customer Name Search", "Column & Operations"], horizontal=True, key="sq_mode")

    if _sq_mode == "Customer Name Search":
        _cust_input = st.text_input("Customer Name", placeholder="e.g. Wipro | Oracle | CISCO", key="sq_cust")
        if st.button("Search", key="sq_run") and _cust_input.strip():
            _pool = combined_df(ALL)
            _cn_col = find_col(_pool, r"customer.*name|client.*name|^customer$")
            if _cn_col:
                _mask = _pool[_cn_col].astype(str).str.lower().str.contains(re.escape(_cust_input.strip().lower()), na=False)
                _found = _pool[_mask]
                if _found.empty:
                    st.warning(f"No rows found for '{_cust_input}'")
                else:
                    st.success(f"Found {len(_found):,} rows across {_found['_Location'].nunique() if '_Location' in _found.columns else 1} locations")
                    st.dataframe(_found, use_container_width=True)
                    st.download_button("Download CSV", _found.to_csv(index=False).encode(), f"customer_{_cust_input[:20]}.csv", "text/csv", key="sq_dl")
    else:
        _op_loc = st.selectbox("Location", ["All"] + sorted(fdata.keys()), key="sq_op_loc")
        _op_pool = CUST.copy()
        if _op_loc != "All" and "_Location" in _op_pool.columns:
            _op_pool = _op_pool[_op_pool["_Location"] == _op_loc]
        _all_c = [c for c in _op_pool.columns if not c.startswith("_")]
        _sel_col = st.selectbox("Column", ["— pick —"] + _all_c, key="sq_col")
        _sel_op = st.selectbox("Operation", ["Sum", "Average", "Min", "Max", "Count", "Median", "Top 10", "Bottom 10", "Show All"], key="sq_op")
        _grp = st.checkbox("Group by Location", value=True, key="sq_grp")
        if st.button("Run", key="sq_run2") and _sel_col != "— pick —":
            _ns = _robust_to_numeric(_op_pool[_sel_col]).dropna()
            if _ns.empty:
                st.warning("No numeric values in this column.")
            else:
                _ops = {"Sum": _ns.sum(), "Average": _ns.mean(), "Min": _ns.min(), "Max": _ns.max(),
                        "Count": float(len(_ns)), "Median": _ns.median()}
                if _sel_op in _ops:
                    _v = _ops[_sel_op]; _u = _detect_unit(_sel_col)
                    st.markdown(f'<div class="result-box"><div class="result-big">{_fmt_decimal(_v)} {_u}</div>'
                        f'<div style="font-size:.8rem;color:{MUTED}">{len(_ns):,} of {len(_op_pool):,} rows</div></div>', unsafe_allow_html=True)
                    if _grp and "_Location" in _op_pool.columns:
                        _gdf = _op_pool.groupby("_Location")[_sel_col].apply(lambda x: _robust_to_numeric(x).sum()).reset_index()
                        _gdf.columns = ["Location", _sel_col]; _gdf = _gdf.sort_values(_sel_col, ascending=False)
                        st.dataframe(_gdf.reset_index(drop=True), use_container_width=True)
                elif _sel_op in ("Top 10", "Bottom 10"):
                    _cn = find_col(_op_pool, r"customer.*name|client.*name")
                    _meta = [c for c in ["_Location", _cn] if c and c in _op_pool.columns]
                    _sub = _op_pool[_meta + [_sel_col]].copy()
                    _sub[_sel_col] = _robust_to_numeric(_sub[_sel_col]); _sub = _sub.dropna(subset=[_sel_col])
                    _sub = _sub.nlargest(10, _sel_col) if "Top" in _sel_op else _sub.nsmallest(10, _sel_col)
                    st.dataframe(_sub.reset_index(drop=True), use_container_width=True)
                else:
                    _meta = [c for c in ["_Location", "_Sheet"] if c in _op_pool.columns]
                    _cn = find_col(_op_pool, r"customer.*name|client.*name")
                    if _cn: _meta.append(_cn)
                    _show = _op_pool[_meta + [_sel_col]].dropna(subset=[_sel_col])
                    st.dataframe(_show.reset_index(drop=True), use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 5 – CROSS-LOCATION
# ══════════════════════════════════════════════════════════════════════════════
with T[5]:
    st.markdown('<div class="section-title">Cross-Location Comparison</div>', unsafe_allow_html=True)
    nc_all = num_cols(CUST)
    if not nc_all:
        st.info("No numeric columns found.")
    else:
        xl1, xl2, xl3 = st.columns(3)
        with xl1: xl_col = st.selectbox("Metric", nc_all, key="xl_col")
        with xl2: xl_op = st.selectbox("Aggregation", ["Sum", "Mean (Avg)", "Max", "Min", "Count"], key="xl_op")
        with xl3: xl_ct = st.selectbox("Chart", ["Bar Chart", "Line Chart", "Box Plot", "Radar Chart"], key="xl_ct")
        rows = []
        for loc in sel_locs:
            loc_df = CUST[CUST["_Location"] == loc] if "_Location" in CUST.columns else CUST
            if not loc_df.empty and xl_col in loc_df.columns:
                val, *_ = run_op(loc_df, xl_col, xl_op)
                if isinstance(val, (int, float)): rows.append({"Location": loc, xl_col: val})
        xl_agg = pd.DataFrame(rows).sort_values(xl_col, ascending=False) if rows else pd.DataFrame()
        if not xl_agg.empty:
            k1, k2, k3 = st.columns(3)
            k1.metric("Highest", xl_agg.iloc[0]["Location"], fmt(xl_agg.iloc[0][xl_col]))
            k2.metric("Lowest", xl_agg.iloc[-1]["Location"], fmt(xl_agg.iloc[-1][xl_col]))
            k3.metric("Network Total", "", fmt(xl_agg[xl_col].sum()))
            if xl_ct == "Radar Chart" and len(xl_agg) >= 3:
                vals = xl_agg[xl_col].tolist(); locs = xl_agg["Location"].tolist()
                fig_xl = go.Figure(go.Scatterpolar(r=vals+[vals[0]], theta=locs+[locs[0]], fill="toself", line_color=CYAN, fillcolor="rgba(0,212,255,0.15)"))
                fig_xl.update_layout(polar=dict(radialaxis=dict(visible=True, gridcolor=BORD), angularaxis=dict(gridcolor=BORD), bgcolor=DARK2), **_base_layout(), height=440)
            else:
                fig_xl = make_chart(xl_ct, xl_agg, "Location", xl_col, title=f"{xl_op} of {xl_col}")
                fig_xl.update_layout(height=420)
            st.plotly_chart(fig_xl, use_container_width=True)
            st.dataframe(xl_agg.round(3), use_container_width=True)
            st.download_button("Download CSV", xl_agg.to_csv(index=False).encode(), "cross_location.csv", "text/csv")
